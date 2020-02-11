# -*- coding: utf-8 -*-

import re
import pandas as pd
import openpyxl,pyperclip,pprint,datetime

# posted = "Posted onMonday, February 03, 2020 12:07:21 AM"
#
# msgtime_regex = re.compile(r'Posted on(.*)')
# ticket_time = msgtime_regex.search(posted).group(1)
#
# time = "Monday, February 03, 2020 12:07:32 AM"
#
# comment_tickets = [['Sunday, February 02, 2020 10:45:40 AM', 'Y L Chan', 'How much?', 'https://www.facebook.com/153753694819713/posts/1249108418617563/?comment_id=1250340885160983'],['Monday, February 03, 2020 11:29:26 AM', 'Chan Wing Yiu', r'https://scontent.xx.fbcdn.net/v/t1.0-9/s720x720/84003939_10157382231709440_3622552945624612864_n.jpg?_nc_cat=102&_nc_oc=AQlq6wf-34Mj-fjynxxUsqUjMdrT_ekpFIBT2Az0-3416A_X124IBKikqCPAeGvuhfY&_nc_ht=scontent.xx&oh=0fa985bb028b65bc4aaf3662ee0bddf4&oe=5EBB5493', r'https://www.facebook.com/153753694819713/posts/1249085878619817/?comment_id=1251058991755839'
# ],['Monday, February 02, 2020 9:29:26 AM', 'Chan Wing Yiu', r'https://scontent.xx.fbcdn.net/v/t1.0-9/s720x720/84003939_10157382231709440_3622552945624612864_n.jpg?_nc_cat=102&_nc_oc=AQlq6wf-34Mj-fjynxxUsqUjMdrT_ekpFIBT2Az0-3416A_X124IBKikqCPAeGvuhfY&_nc_ht=scontent.xx&oh=0fa985bb028b65bc4aaf3662ee0bddf4&oe=5EBB5493', r'https://www.facebook.com/153753694819713/posts/1249085878619817/?comment_id=1251058991755839']]
#
# inbox_tickets = [['Monday, February 03, 2020 12:07:32 AM', 'HoWai Lam', '門市有得買？', 'Inbox'], ['Monday, February 03, 2020 12:07:21 AM', 'HoWai Lam', 'https://s3.amazonaws.com/spr-uploads-prod/815/FACEBOOK/82906/media/192A7B70A2A4B7997448C1AB747EE185', 'Inbox']]
#
# df_comment = pd.DataFrame(comment_tickets,columns=["Message_Time","Customer Name","Enquiry","URL"])
#
# df_inbox = pd.DataFrame(inbox_tickets,columns=["Message Time","Customer Name","Enquiry","URL"])
#
# df_comment["Message_Time"] = pd.to_datetime(df_comment.Message_Time)
#
# df_comment = df_comment.sort_values(by ='Message_Time',ascending=False)
#
# df_comment["Message_Time"] = df_comment["Message_Time"].dt.strftime("%A, %B %d, %Y %I:%M:%S %p")
#
# """print(x.strftime("%b %d %Y %H:%M:%S"))
# Output:
#
# Sep 15 2018 00:00:00"""
#
# print(df_comment)

def delete_sheets():

    export_sheet = ["Comment", "Inbox"]
    wb_path = r'C:\Users\abc\PycharmProjects\somestuff\pandora_scraper\pandora_export.xlsx'
    wb = openpyxl.load_workbook(wb_path)
    sheets = wb.sheetnames
    for i in export_sheet:
        if i in sheets:
            wb.remove(wb[i])
    wb.save(wb_path)

def df_to_list():
    df = pd.DataFrame({'a': [1, 2, 3, 4],
                       'b': [5, 6, 7, 8]})

    print(df)
    print(df.iterrows())
    liss = []
    for index,row in df.iterrows():
        a = [row["a"],row["b"]]
        liss.append(a)
        print(a)
    print(liss)
    #print('list', [df[i].tolist() for i in df.iterrows()])

def test_list():
    tickets_needed_to_reply = [
        ['Wednesday, February 05, 2020 3:12:30 AM', 'Lok Daisy', '拎條鍊去洗，會屈你條鍊用洗銀水浸過，然後同你講洗唔到原來既色架！仲買？',
         'https://www.facebook.com/153753694819713/posts/1251826478345757/?comment_id=1252278748300530'],
        ['Wednesday, February 05, 2020 7:49:58 AM', 'Crystal Man', '價錢超貴，質量極差👎仲衰過爛銅爛鐵😤',
         'https://www.facebook.com/153753694819713/posts/1251251691736569/?comment_id=1252428578285547'],
    ]

    responses = [
        tickets_needed_to_reply[0] + ["complaints", "No response required"],
        tickets_needed_to_reply[1] + ["complaints", "No response required"]
    ]

    print(responses)


#reply, tag = get_reponse_for_one_ticket(ticket)
#print((reply,tag))


#
# tickets_to_re = """[[['Thursday, February 06, 2020 14:28:07', 'Helen Helen', '我都覺得呢個牌子唔應該再買，因脱色好緊要。', 'https://www.facebook.com/153753694819713/posts/1249085878619817/?comment_id=1253319778196427'], ['Hi
# Helen Helen，非常抱歉你有以上經歷。請你電郵至 pap.cs@Pandora.net 提供姓名、聯絡電話同詳情，我哋會作出跟進。多謝！', 'Complaints', '']], [['Thursday, February 06, 2020 14:35:47', 'Keith Hang', '口罩漂白水酒
# 精\n溝死女\n駛Q買你d珠仔💪🏻', 'https://www.facebook.com/153753694819713/posts/1251808008347604/?comment_id=1253322551529483'], ['No response required', 'Other', '']], [['Thursday, February 06, 2020 16:
# 42:38', 'Lok Daisy', '拎條鍊去洗，會屈你條鍊用洗銀水浸過，然後同你講洗唔到原來既色架！仲買？', 'https://www.facebook.com/153753694819713/posts/2240275839408885/?comment_id=2240277299408739'], ['No response
# required', 'Complaints', '']], [['Thursday, February 06, 2020 18:58:27', 'Beau Manuswee', 'Gift Kumpunthong อิอิ เพื่อคันๆๆๆ', 'https://www.facebook.com/153753694819713/posts/1249085878619817/?comment_id=12
# 53444781517260'], ['No response required', 'Other', '']], [['Thursday, February 06, 2020 21:43:47', 'Lok Daisy', '拎條鍊去洗，會屈你條鍊用洗銀水浸過，然後同你講洗唔到原來既色架！仲買？', 'https://www.facebo
# ok.com/153753694819713/posts/2790696994324388/?comment_id=2790697870990967'], ['No response required', 'Complaints', '']], [['Thursday, February 06, 2020 23:28:10', 'May Yuen', '請問米奇要幾錢', 'https://ww
# w.facebook.com/153753694819713/posts/1242183959310009/?comment_id=1253608004834271'], ['Hi May Yuen，請到 https://go.pandora.net/2SqrqAj 查閱價錢，多謝你嘅支持！', 'General', '']], [['Friday, February 07, 2
# 020 08:31:24', 'Mavis Law', 'Koey Wai Cherry Yeung', 'https://www.facebook.com/153753694819713/posts/1251251691736569/?comment_id=1253890828139322'], ['No response required', 'Other', '']], [['Thursday, Feb
# ruary 06, 2020 14:33:51', 'Kathy Ching', "i wanted to ask if there's stock of harry potter series in kings cross shop.  i have purchased on Tuesday already, thx", 'Inbox'], ['Hi Kathy Ching, thanks for supp
# orting us! Love Pandora!', 'Other', '']], [['Friday, February 07, 2020 12:03:19', 'Fanny Tam', '買滿幾錢？', 'https://www.facebook.com/153753694819713/posts/1251251691736569/?comment_id=1253995441462194'],
# ['Hi Fanny Tam，而家喺網店同專門店消費淨值滿港幣$1,380，即可獲贈指定項鏈；\n淨值滿港幣$2,280，即可獲贈指定項鏈及耳環。優惠去到2020年2月15日 00:00，贈品數量有限，送完即止。多謝支持Pandora！', 'General', '']]
# , [['Friday, February 07, 2020 13:49:29', 'Keong Keong', '就像愛情，找帥哥破處X膜，然後被拋棄，留下永久撕裂傷，開始放盪，在去做處X膜手術，然後找好騙工具人，一帥哥渣男多處X制，聽歌聽到腦殘，該感動不感動，亂
# 享受感動，越漂亮曾墮胎機率越高👺👺👺👹👹👹', 'https://www.facebook.com/PandoraHongKong/photos/a.156393951222354/1251221261739612/?type=3&comment_id=1254066051455133'], ['Hide this comment, mark
# as spam', 'Other', '']], [['Friday, February 07, 2020 14:23:06', 'Cherry Yeung', 'Mavis Law 咁要集結去睇喎🤣', 'https://www.facebook.com/153753694819713/posts/1251251691736569/?comment_id=1253890828139322
# &reply_comment_id=1254083041453434'], ['No response required', 'Other', '']], [['Friday, February 07, 2020 13:16:07', 'Typ Wong', 'https://ezone.ulifestyle.com.hk/article/2557002', 'Inbox'], ['No response r
# equired', 'Other', '']], [['Friday, February 07, 2020 14:42:50', 'Lillian Li', '我門市買夠數，有9折，已經試過起門市用到9折，但網上又話我電話沒有登記，請跟進！\n\nTel: 62906286', 'Inbox'], ['Hi Lillian Li，
# 非常抱歉你有以上經歷。請你電郵至 pap.cs@Pandora.net 提供姓名、聯絡電話同詳情，我哋會作出跟進。多謝你嘅支持！', 'General', '']], [['Friday, February 07, 2020 14:43:06', 'Lillian Li', 'https://s3.amazonaws.co
# m/spr-uploads-prod/815/FACEBOOK/82906/media/F288E0AB9EF90AFF734CCA123FC44B2A', 'Inbox'], ['No response required', 'General', '']]]
# """
# tickets_to_re = [[['Thursday, February 06, 2020 14:28:07', 'Helen Helen', '我都覺得呢個牌子唔應該再買，因脱色好緊要。', 'https://www.facebook.com/153753694819713/posts/1249085878619817/?comment_id=1253319778196427'], ['HiHelen Helen，非常抱歉你有以上經歷。請你電郵至 pap.cs@Pandora.net 提供姓名、聯絡電話同詳情，我哋會作出跟進。多謝！', 'Complaints', '']], [['Thursday, February 06, 2020 14:35:47', 'Keith Hang', '口罩漂白水酒精溝死女駛Q買你d珠仔💪🏻', 'https://www.facebook.com/153753694819713/posts/1251808008347604/?comment_id=1253322551529483'], ['No response required', 'Other', '']], [['Thursday, February 06, 2020 16:42:38', 'Lok Daisy', '拎條鍊去洗，會屈你條鍊用洗銀水浸過，然後同你講洗唔到原來既色架！仲買？', 'https://www.facebook.com/153753694819713/posts/2240275839408885/?comment_id=2240277299408739'], ['No responserequired', 'Complaints', '']], [['Thursday, February 06, 2020 18:58:27', 'Beau Manuswee', 'Gift Kumpunthong อิอิ เพื่อคันๆๆๆ', 'https://www.facebook.com/153753694819713/posts/1249085878619817/?comment_id=1253444781517260'], ['No response required', 'Other', '']], [['Thursday, February 06, 2020 21:43:47', 'Lok Daisy', '拎條鍊去洗，會屈你條鍊用洗銀水浸過，然後同你講洗唔到原來既色架！仲買？', 'https://www.facebook.com/153753694819713/posts/2790696994324388/?comment_id=2790697870990967'], ['No response required', 'Complaints', '']], [['Thursday, February 06, 2020 23:28:10', 'May Yuen', '請問米奇要幾錢', 'https://www.facebook.com/153753694819713/posts/1242183959310009/?comment_id=1253608004834271'], ['Hi May Yuen，請到 https://go.pandora.net/2SqrqAj 查閱價錢，多謝你嘅支持！', 'General', '']], [['Friday, February 07, 2020 08:31:24', 'Mavis Law', 'Koey Wai Cherry Yeung', 'https://www.facebook.com/153753694819713/posts/1251251691736569/?comment_id=1253890828139322'], ['No response required', 'Other', '']], [['Thursday, February 06, 2020 14:33:51', 'Kathy Ching', "i wanted to ask if there's stock of harry potter series in kings cross shop.  i have purchased on Tuesday already, thx", 'Inbox'], ['Hi Kathy Ching, thanks for supporting us! Love Pandora!', 'Other', '']], [['Friday, February 07, 2020 12:03:19', 'Fanny Tam', '買滿幾錢？', 'https://www.facebook.com/153753694819713/posts/1251251691736569/?comment_id=1253995441462194'],['Hi Fanny Tam，而家喺網店同專門店消費淨值滿港幣$1,380，即可獲贈指定項鏈；淨值滿港幣$2,280，即可獲贈指定項鏈及耳環。優惠去到2020年2月15日 00:00，贈品數量有限，送完即止。多謝支持Pandora！', 'General', '']], [['Friday, February 07, 2020 13:49:29', 'Keong Keong', '就像愛情，找帥哥破處X膜，然後被拋棄，留下永久撕裂傷，開始放盪，在去做處X膜手術，然後找好騙工具人，一帥哥渣男多處X制，聽歌聽到腦殘，該感動不感動，亂享受感動，越漂亮曾墮胎機率越高👺👺👺👹👹👹', 'https://www.facebook.com/PandoraHongKong/photos/a.156393951222354/1251221261739612/?type=3&comment_id=1254066051455133'], ['Hide this comment, markas spam', 'Other', '']], [['Friday, February 07, 2020 14:23:06', 'Cherry Yeung', 'Mavis Law 咁要集結去睇喎🤣', 'https://www.facebook.com/153753694819713/posts/1251251691736569/?comment_id=1253890828139322&reply_comment_id=1254083041453434'], ['No response required', 'Other', '']], [['Friday, February 07, 2020 13:16:07', 'Typ Wong', 'https://ezone.ulifestyle.com.hk/article/2557002', 'Inbox'], ['No response required', 'Other', '']], [['Friday, February 07, 2020 14:42:50', 'Lillian Li', '我門市買夠數，有9折，已經試過起門市用到9折，但網上又話我電話沒有登記，請跟進！Tel: 62906286', 'Inbox'], ['Hi Lillian Li，非常抱歉你有以上經歷。請你電郵至 pap.cs@Pandora.net 提供姓名、聯絡電話同詳情，我哋會作出跟進。多謝你嘅支持！', 'General', '']], [['Friday, February 07, 2020 14:43:06', 'Lillian Li', 'https://s3.amazonaws.com/spr-uploads-prod/815/FACEBOOK/82906/media/F288E0AB9EF90AFF734CCA123FC44B2A', 'Inbox'], ['No response required', 'General', '']]]
# pprint.pprint(tickets_to_re)



def responses():
    excel_path = r'C:\Users\abc\PycharmProjects\somestuff\pandora_scraper\pandora_export.xlsx'
    wb = openpyxl.load_workbook(excel_path,data_only=True)
    sheet = wb["response"]
    sheet_max_row = sheet.max_row

    # No. Message Time	Customer Name  Enquiry  Reply(modified) URL	Category/Tag	Comment from PANDORA	KathyReply	replyOrNot
    tickets_and_responses = []
    for eachRow in range(2,sheet_max_row+1):
        ticket_detail_and_reponse = [[],[]]

        for specificCol in [2,3,4,6]:
            ticket_spec = sheet.cell(row=eachRow,column=specificCol).value
            if ticket_spec == 0 or ticket_spec == None:
                ticket_spec = ""
            ticket_detail_and_reponse[0].append(ticket_spec)

        for ticket_response_col in [5,7,10]:
            ticket_responses = sheet.cell(row=eachRow,column=ticket_response_col).value
            if ticket_responses == 0 or ticket_responses == None:
                ticket_responses = ""
            ticket_detail_and_reponse[1].append(ticket_responses)

        the_ticket_time = ticket_detail_and_reponse[0][0]

        """oh"""


        reply_or_not = ticket_detail_and_reponse[1][2]
        if the_ticket_time != "" and reply_or_not !="n":
            tickets_and_responses.append(ticket_detail_and_reponse)

    print(tickets_and_responses)
    return  tickets_and_responses

ticket = ['Tuesday, February 04, 2020 19:26:56', 'Djq Wong', '垃圾', 'https://www.facebook.com/153753694819713/posts/1249085878619817/?comment_id=1252008424994229']
def get_reponse_for_one_ticket(ticket):
    tickets_and_responses = responses()
    for ticket_and_response in tickets_and_responses:
        if ticket in ticket_and_response:
            response = ticket_and_response[1]
            reply = response[0]
            tag = response[1]
            return reply, tag

# right format "%A, %B %d, %Y %I:%M:%S %p"
time_a = "Thursday, February 06, 2020 15:28:07"
date_obj = datetime.datetime.strptime(time_a,'%A, %B %d, %Y %H:%M:%S')
time_b = date_obj.strftime('%A, %B %d, %Y ')
time_bb = date_obj.strftime('%I:%M:%S %p').lstrip('0')
time_c = time_b + time_bb
print(time_c)
print(time_bb)
print(time_b)


#python -i pandora_scraper.py
# a.rtest(scrolltimes=70,c_or_i="inbox")
# ['Monday, February 10, 2020 5:37:48 AM', 'Dave Henz', 'https://s3.amazonaws.com/spr-uploads-prod/815/FACEBOOK/82906/media/F279C2ECD2CC380F0E54A0B4103F302C', 'Inbox']
# ['Monday, February 10, 2020 5:37:48 PM', 'Dave Henz', 'https://s3.amazonaws.com/spr-uploads-prod/815/FACEBOOK/82906/media/F279C2ECD2CC380F0E54A0B4103F302C', 'Inbox']

def tt():

    t = ['Friday, February 07, 2020 1:49:29 PM', 'Keong Keong', '就像愛情，找帥哥破處X膜，然後被拋棄，留下永久撕裂傷，開始放盪，在去做處X膜手術，然後找好騙工具人，一帥哥渣男多處X制，聽歌聽到腦殘，該感動不感動，亂享受感動，越漂亮曾墮胎機率越高👺👺👺👹👹👹', 'https://www.facebook.com/PandoraHongKong/photos/a.156393951222354/1251221261739612/?type=3&comment_id=1254066051455133']
    r =  [['Friday, February 07, 2020 01:49:29 PM', 'Keong Keong', '就像愛情，找帥哥破處X膜，然後被拋棄，留下永久撕裂傷，開始放盪，在去做處X膜手術，然後找好騙工具人，一帥哥渣男多處X制，聽歌聽到腦殘，該感動不感動，亂享受感動，越漂亮曾墮胎機率越高👺👺👺👹👹👹', 'https://www.facebook.com/PandoraHongKong/photos/a.156393951222354/1251221261739612/?type=3&comment_id=1254066051455133'], ['Hide this comment,mark as spam', 'Other', '']]
