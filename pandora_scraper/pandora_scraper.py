#  -*- coding: utf-8 -*-

from selenium import webdriver
import time, re,datetime
from secret import username, password
import pandas as pd
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import pyautogui

#python -i pandora_scraper.py

class pandoraBot:

    comment_tickets = []
    inbox_tickets = []

    tickets_needed_to_reply = []

    export_colomns = ["Message_Time","Customer_Name","Enquiry","URL"]

    export_path = r"C:\Users\abc\PycharmProjects\somestuff\pandora_scraper\pandora_export.xlsx"

    def __init__(self):
        self.driver = webdriver.Chrome(executable_path=r"C:\Users\abc\PycharmProjects\chromedriver_win32\chromedriver.exe")

    def _login_old(self):
        self.driver.get("https://space.sprinklr.com/new")

        time.sleep(7)

        email_input = self.driver.find_element_by_xpath('//*[@id="root"]/div/section[2]/section/div[1]/div/div/div/div/div/div/form/div[1]/div/input')
        time.sleep(0.5)
        email_input.send_keys(username)
        time.sleep(1)

        next_btn = self.driver.find_element_by_xpath('//*[@id="root"]/div/section[2]/section/div[1]/div/div/div/div/div/div/form/div[1]/button')
        time.sleep(0.5)
        next_btn.click()
        time.sleep(1)

        pw_input = self.driver.find_element_by_xpath('//*[@id="root"]/div/section[2]/section/div[1]/div/div/div/div/div/div[2]/form/div[2]/div[1]/input')
        time.sleep(0.5)
        pw_input.send_keys(password)
        time.sleep(1)

        login_btn = self.driver.find_element_by_xpath('//*[@id="root"]/div/section[2]/section/div[1]/div/div/div/div/div/div[2]/form/div[2]/button')
        time.sleep(0.5)
        login_btn.click()

        time.sleep(12)

        engagement_open = self.driver.find_element_by_xpath('//*[@id="sprPage"]/div/div/div/div/div[2]/div/div/div/div/div/ul/li[1]/div[2]/div/div/div[1]/a[1]/div/div/div[1]')
        time.sleep(0.5)
        engagement_open.click()
        time.sleep(2)

    def login(self):
        self.driver.get("https://space.sprinklr.com/new")

        time.sleep(2)

        email_input_element_not_found = True
        while email_input_element_not_found:
            try:
                email_input = self.driver.find_element_by_xpath('//*[@id="root"]/div/section[2]/section/div[1]/div/div/div/div/div/div/form/div[1]/div/input')
                email_input_element_not_found = False
                print("email input element found")
            except:
                print("retry finding email input element...")
            time.sleep(0.5)

        email_input.send_keys(username)
        time.sleep(1)

        next_btn = self.driver.find_element_by_xpath(
            '//*[@id="root"]/div/section[2]/section/div[1]/div/div/div/div/div/div/form/div[1]/button')
        time.sleep(0.5)
        next_btn.click()
        time.sleep(1)

        pw_input_element_not_found = True
        while pw_input_element_not_found:
            try:
                pw_input = self.driver.find_element_by_xpath(
                    '//*[@id="root"]/div/section[2]/section/div[1]/div/div/div/div/div/div[2]/form/div[2]/div[1]/input')
                print("pw input element found")
                pw_input_element_not_found = False
            except:
                print("retry finding pw input element")
            time.sleep(0.5)

        pw_input.send_keys(password)
        time.sleep(1)

        login_btn = self.driver.find_element_by_xpath(
            '//*[@id="root"]/div/section[2]/section/div[1]/div/div/div/div/div/div[2]/form/div[2]/button')
        time.sleep(0.5)
        login_btn.click()

        time.sleep(2)

        engagement_open_element_not_found = True
        while engagement_open_element_not_found:
            try:
                engagement_open = self.driver.find_element_by_xpath(
                    '//*[@id="sprPage"]/div/div/div/div/div[2]/div/div/div/div/div/ul/li[1]/div[2]/div/div/div[1]/a[1]/div/div/div[1]')
                engagement_open_element_not_found = False
                print("engagement_open_element found")
            except:
                print("retry finding engagement_open_element")
            time.sleep(0.5)

        engagement_open.click()
        time.sleep(2)

    def getTicketContent(self):

        ticket_ele = self.driver.find_elements_by_xpath('//article[@class = "streamEntity spr"]')

        comment_tickets = []
        inbox_tickets = []

        for i in ticket_ele:
            ticket_name = i.find_element_by_xpath('.//div[@class="msgProfileName spr-text-01 text-13 font-700 text-overflow scp"]').text
            ticket_text = i.find_element_by_xpath('.//div[@class="msgContent"]').text

            try:
                ticket_link = i.find_element_by_xpath('.//a[@class = "msgTimeStamp msgHeaderSubText spr-text-02 txt-bd4 pull-xs-right m-l-1 msgTimeStampRenderAsLink"]').get_attribute("href")
                if "inbox" in ticket_link:
                    ticket_link = "Inbox"
            except:
                ticket_link = ""

            try:
                ticket_time = i.find_element_by_xpath('.//a[contains(@aria-label,"Posted on")]').get_attribute("aria-label")
                msgtime_regex = re.compile(r'Posted on(.*)')
                if ticket_time:
                    ticket_time = msgtime_regex.search(ticket_time).group(1)
            except:
                ticket_time = ""

            try:
                ticket_image = i.find_element_by_xpath('.//img[@class = "show mediaImgContent scp"]').get_attribute("src")
            except:
                ticket_image = None

            if ticket_image and ticket_text:
                ticket_text = ticket_text + "\n" + ticket_image
            elif ticket_image and ticket_text == "":
                ticket_text = ticket_image

            per_ticket_content = [ticket_time,ticket_name,ticket_text,ticket_link]
            if ticket_link == "Inbox":
                if per_ticket_content not in self.inbox_tickets:
                    self.inbox_tickets.append(per_ticket_content)
            elif ticket_link != "" and ticket_link != "Inbox":
                if per_ticket_content not in self.comment_tickets:
                    self.comment_tickets.append(per_ticket_content)


            #print(ticket_name,ticket_text,ticket_link,ticket_time,"\n")
            # if ticket_image:
            #     print(ticket_image,"\n")

    def rtest(self,scrolltimes=40,c_or_i = "comment"):

        tickets_and_responses = self._responses()
        self.tickets_needed_to_reply = tickets_and_responses

        # responses = [
        #     tickets_needed_to_reply[0] + ["complaints","No response required"],
        #     tickets_needed_to_reply[1] + ["other","No response required"]
        # ]

        scroll_box_public = self.driver.find_element_by_xpath(
            '//*[@id="sprEngagementWorkspace"]/div/div/div[2]/div/div[2]/div[2]/div/div/section/div')

        scroll_box_private = self.driver.find_elements_by_xpath('//*[@id="sprEngagementWorkspace"]/div/div/div[3]/div/div[2]/div[2]/div/div/section/div')

        scroll_height = -30
        for i in range(scrolltimes):
            scroll_height += 30
            if c_or_i == "comment":
                self.driver.execute_script("arguments[0].scrollTo(0, {})".format(scroll_height), scroll_box_public)
            elif c_or_i == "inbox":
                self.driver.execute_script("arguments[0].scrollTo(0, {})".format(scroll_height), scroll_box_private[0])
            time.sleep(0.2)
            ticket_frame = self.driver.find_element_by_xpath(
                '//*[@id="sprEngagementWorkspace"]/div/div/div[2]/div/div[2]/div[2]/div/div/section/div/div/div/div/article[1]')
            ticket_name = ticket_frame.find_element_by_xpath(
                './/div[@class="msgProfileName spr-text-01 text-13 font-700 text-overflow scp"]').text
            ticket_text = ticket_frame.find_element_by_xpath('.//div[@class="msgContent"]').text

            try:
                ticket_link = ticket_frame.find_element_by_xpath('.//a[@class = "msgTimeStamp msgHeaderSubText spr-text-02 txt-bd4 pull-xs-right m-l-1 msgTimeStampRenderAsLink"]').get_attribute("href")
                if "inbox" in ticket_link:
                    ticket_link = "Inbox"
            except:
                ticket_link = ""

            try:
                ticket_time = ticket_frame.find_element_by_xpath('.//a[contains(@aria-label,"Posted on")]').get_attribute("aria-label")
                msgtime_regex = re.compile(r'Posted on(.*)')
                if ticket_time:
                    ticket_time = msgtime_regex.search(ticket_time).group(1)
            except:
                ticket_time = ""

            try:
                ticket_image = ticket_frame.find_element_by_xpath('.//img[@class = "show mediaImgContent scp"]').get_attribute("src")
            except:
                ticket_image = None

            if ticket_image and ticket_text:
                ticket_text = ticket_text + "\n" + ticket_image
            elif ticket_image and ticket_text == "":
                ticket_text = ticket_image

            per_ticket_content = [ticket_time, ticket_name, ticket_text, ticket_link]

            print(per_ticket_content)

            reply, tag = self.get_response_for_one_ticket(self.tickets_needed_to_reply,per_ticket_content)
            print("reply:",reply,"tag:",tag)

            #if per_ticket_content in tickets_needed_to_reply:
            if reply != "" or tag != "":
                print(reply,tag)

                ticket_frame.click()
                time.sleep(1)

                for i in self.tickets_needed_to_reply:
                    the_ticket = i[0]
                    if the_ticket == per_ticket_content:
                        self.tickets_needed_to_reply.remove(i)

                wait = WebDriverWait(self.driver, 10)

                actions = ActionChains(self.driver)
                actions.move_to_element(ticket_frame)
                actions.perform()
                print("hovering on the ticket")

                #reply_btn = wait.until(EC.visibility_of_element_located((By.XPATH,'//*[@id="sprEngagementWorkspace"]/div/div/div[2]/div/div[2]/div[2]/div/div/section/div/div/div/div/article[1]/section/section/article/div/div/section[2]/div[2]/div[3]/div/span/span[3]/button')))
                time.sleep(1)

                if reply != "No response required" and reply != "":

                    print("have to reply this")

                    for i in range(8):
                        replyActions = ActionChains(self.driver)
                        replyActions = replyActions.send_keys(Keys.TAB)
                        replyActions.perform()
                        time.sleep(0.7)
                        try:
                            reply_tip = self.driver.find_element_by_xpath(
                                '//div[@class="tooltip-inner" and contains(text(),"Reply")]')
                            print("reply tip found")
                            actions = ActionChains(self.driver)
                            actions = actions.send_keys(Keys.SPACE)
                            actions.perform()
                            print("opening reply box")
                            break
                        except:
                            pass
                    print("reply box opened")

                    reply_box_input = wait.until(EC.visibility_of_element_located((By.XPATH,
                                                                                   '//*[@id="sprBasePublisher_REPLY"]/div/section[3]/div/div/div[1]/div/div/div/div/div[2]/div')))
                    reply_box_input.send_keys(reply)
                    time.sleep(0.5)

                    confirming_reply = pyautogui.confirm(text="I'm gonna reply this",title="Confirming",buttons=["OK","Noooo"])
                    if confirming_reply == "OK":
                        post_btn = self.driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/article/div[3]/div[2]/div/button[2]')
                        self.driver.execute_script("arguments[0].click();", post_btn)
                        time.sleep(1.5)

                    time.sleep(0.8)
                    close_reply_box_btn = self.driver.find_element_by_xpath('//div[@class="-kp0 spr p-l-5 p-x-3"]/button[@type = "button" and @class = "iconButton fill-icon-button center-x-y icon-btn-cont-30 focus-border"]')
                    self.driver.execute_script("arguments[0].click();", close_reply_box_btn)
                    print("reply box closed")
                    time.sleep(0.5)

                if tag != "":

                    print("have to update tag")

                    """update tags"""
                    dot_batch_ele = wait.until(EC.visibility_of_element_located((By.XPATH,'//*[@id="sprBody"]/section/div[2]/div[1]/article/section[2]/section/div/div[5]')))
                    actions = ActionChains(self.driver)
                    actions.move_to_element(dot_batch_ele).perform()

                    update_tag_btn = wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@class="_1ZpM flex-item-1" and contains(text(),"Update Tags")]')))
                    update_tag_btn.click()

                    tag_arrow = wait.until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[4]/div/div[2]/div[2]/div[1]/div/div/div/span[2]')))
                    tag_arrow.click()

                    tag_select = wait.until(EC.visibility_of_element_located((By.XPATH,'//input[@role="combobox"]')))
                    tag_select.send_keys(tag)
                    tag_select.send_keys(Keys.ENTER)
                    time.sleep(0.5)

                    update_btn = self.driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/div[3]/div/button[2]')
                    update_btn.click()

                time.sleep(1)

        print(self.tickets_needed_to_reply)

    def get_dataframe(self):

        df_comment = pd.DataFrame(columns=self.export_colomns)
        df_inbox = pd.DataFrame(columns=self.export_colomns)

        if len(self.comment_tickets) > 0:
            df_comment = pd.DataFrame(self.comment_tickets,columns=self.export_colomns)

            df_comment["Message_Time"] = pd.to_datetime(df_comment.Message_Time)
            df_comment = df_comment.sort_values(by='Message_Time', ascending=True)

            df_comment["Message_Time"] = df_comment["Message_Time"].dt.strftime("%A, %B %d, %Y %I:%M:%S %p")

            print(df_comment)
            print(self.comment_tickets)
        if len(self.inbox_tickets) > 0:
            df_inbox = pd.DataFrame(self.inbox_tickets,columns=self.export_colomns)

            df_inbox["Message_Time"] = pd.to_datetime(df_inbox.Message_Time)
            df_inbox = df_inbox.sort_values(by='Message_Time',ascending=True)

            df_inbox["Message_Time"] = df_inbox["Message_Time"].dt.strftime("%A, %B %d, %Y %I:%M:%S %p")

            print(df_inbox)
            print(self.inbox_tickets)

        return df_comment,df_inbox

    def getTickets(self,scrolls = 5 ):
        scroll_box_public = self.driver.find_element_by_xpath('//*[@id="sprEngagementWorkspace"]/div/div/div[2]/div/div[2]/div[2]/div/div/section/div')
        scroll_box_private = self.driver.find_elements_by_xpath('//*[@id="sprEngagementWorkspace"]/div/div/div[3]/div/div[2]/div[2]/div/div/section/div')
        #scroll_box_private[0].location_once_scrolled_into_view
        #scroll_box_public[0].location_once_scrolled_into_view
        # random_user_private = a.driver.find_elements_by_xpath('//*[@id="sprEngagementWorkspace"]/div/div/div[3]/div/div[2]/div[2]/div/div/section/div/div/div/div/article[2]/section/section/article/div/div/section[2]/div[1]/div/div[1]/div[1]/div/div')
        # random_user_private = a.driver.find_elements_by_xpath('//div[contains(text(),"Bleu Loi")]')
        # random_user_private[0].location_once_scrolled_into_view

        # from selenium.webdriver.common.action_chains import ActionChains
        # ActionChains(a.driver).move_to_element(scroll_box_private).perform()

        scroll_height = -260
        for _ in range(scrolls):
            scroll_height += 260
            self.driver.execute_script("arguments[0].scrollTo(0, {})".format(scroll_height), scroll_box_public)
            time.sleep(0.2)
            #a.driver.execute_script("arguments[0].scrollTo(0, {})".format(scroll_height), scroll_box_private)
            self.driver.execute_script("arguments[0].scrollTo(0, {})".format(scroll_height),scroll_box_private[0])
            time.sleep(0.2)
            self.getTicketContent()
        self.get_dataframe()

    def _responses(self):
        wb = openpyxl.load_workbook(self.export_path, data_only=True)
        sheet = wb["response"]
        sheet_max_row = sheet.max_row

        # No. Message Time	Customer Name  Enquiry  Reply(modified) URL	Category/Tag	Comment from PANDORA	KathyReply	replyOrNot
        tickets_and_responses = []
        for eachRow in range(2, sheet_max_row + 1):
            ticket_detail_and_response = [[], []]

            for specificCol in [2, 3, 4, 6]:
                ticket_spec = sheet.cell(row=eachRow, column=specificCol).value
                if ticket_spec == 0 or ticket_spec == None:
                    ticket_spec = ""
                if specificCol == 2 and ticket_spec != "":
                    date_obj = datetime.datetime.strptime(ticket_spec, '%A, %B %d, %Y %H:%M:%S')
                    time_a = date_obj.strftime('%A, %B %d, %Y ')
                    time_b = date_obj.strftime('%I:%M:%S %p').lstrip('0')
                    ticket_spec = time_a + time_b
                ticket_detail_and_response[0].append(ticket_spec)

            for ticket_response_col in [5, 7, 10]:
                ticket_responses = sheet.cell(row=eachRow, column=ticket_response_col).value
                if ticket_responses == 0 or ticket_responses == None:
                    ticket_responses = ""
                ticket_detail_and_response[1].append(ticket_responses)

            the_ticket_time = ticket_detail_and_response[0][0]
            reply_or_not = ticket_detail_and_response[1][2]
            if the_ticket_time != "" and reply_or_not != "n" and reply_or_not !="Y":
                tickets_and_responses.append(ticket_detail_and_response)

        print(tickets_and_responses)
        return tickets_and_responses

    def get_response_for_one_ticket(self,tickets_and_responses,ticket):
        reply = ""
        tag = ""
        for ticket_and_response in tickets_and_responses:
            if ticket in ticket_and_response:
                response = ticket_and_response[1]
                reply = response[0]
                tag = response[1]
                break
        return reply,tag

    def export_to_excel(self):

        export_sheet = ["Comment", "Inbox"]

        # delete the previous export sheets
        wb = openpyxl.load_workbook(self.export_path)
        sheets = wb.sheetnames
        for i in export_sheet:
            if i in sheets:
                wb.remove(wb[i])
        wb.save(self.export_path)


        df_comment, df_inbox = self.get_dataframe()

        dfs = [df_comment,df_inbox]

        for i,df in enumerate(dfs):
            wb = openpyxl.load_workbook(self.export_path)
            writer = pd.ExcelWriter(self.export_path, engine='openpyxl')
            writer.book = wb

            df.to_excel(writer, sheet_name=export_sheet[i], index=False)
            writer.save()
            writer.close()

    def export_with_openpyxl(self):

        #per_ticket_content = [ticket_time,ticket_name,ticket_text,ticket_link]

        df_comment, df_inbox = self.get_dataframe()

        if len(self.comment_tickets) > 0:

            ordered_comment_tickets = []
            for index, row in df_comment.iterrows():
                ticket = [row[self.export_colomns[0]],row[self.export_colomns[1]],row[self.export_colomns[2]],row[self.export_colomns[3]]]
                ordered_comment_tickets.append(ticket)

            wb = openpyxl.load_workbook(self.export_path)
            sheet_public = wb["Public"]

            sheet_public.delete_rows(2, 200)

            for i,each_ticket in enumerate(ordered_comment_tickets):
                for n,each_piece_of_info in enumerate(each_ticket):
                    sheet_public.cell(row=2+i,column=n+1).value = each_piece_of_info

            wb.save(self.export_path)

        if len(self.inbox_tickets) > 0:

            ordered_inbox_tickets = []
            for index, row in df_inbox.iterrows():
                ticket = [row[self.export_colomns[0]],row[self.export_colomns[1]],row[self.export_colomns[2]],row[self.export_colomns[3]]]
                ordered_inbox_tickets.append(ticket)

            wb = openpyxl.load_workbook(self.export_path)
            sheet_private = wb["Private"]

            sheet_private.delete_rows(2,200)

            for i, each_ticket in enumerate(ordered_inbox_tickets):
                for n, each_piece_of_info in enumerate(each_ticket):
                    sheet_private.cell(row=2+i,column=n+1).value = each_piece_of_info

            wb.save(self.export_path)

    def _reply(self):
        ticket_frame = a.driver.find_element_by_xpath('//*[@id="sprEngagementWorkspace"]/div/div/div[2]/div/div[2]/div[2]/div/div/section/div/div/div/div/article[1]')
        specific_user = a.driver.find_element_by_xpath('//div[contains(text(),"Kit Derek Kam")]')
        specific_content = a.driver.find_element_by_xpath('//span[contains(text(),"親生果个都無，你愛黎做乜")]')
        hover = ActionChains(a.driver).move_to_element(ticket_frame)
        hover.perform()
        reply_btn = a.driver.find_element_by_xpath('//*[@id="sprEngagementWorkspace"]/div/div/div[2]/div/div[2]/div[2]/div/div/section/div/div/div/div/article[1]/section/section/article/div/div/section[2]/div[2]/div[3]/div/span/span[3]/button')
        reply_btn.click()

        reply_box = a.driver.find_element_by_xpath('//*[@id="sprBasePublisher_REPLY"]/div/section[3]/div/div/div[1]/div/div/div/div/div[2]/div')
        reply_box.send_keys("test")

        close_reply_box_btn = a.driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/article/div[1]/button')
        close_reply_box_btn.click()

        post_btn = a.driver.find_element_by_xpath('/html/body/div[4]/div/div[2]/article/div[3]/div[2]/div/button[2]')
        post_btn.click()





a = pandoraBot()
a.login()





